import csv
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from .base import BillParseError


def _csv_rows(path: Path) -> Iterator[list[str]]:
    raw = path.read_bytes()
    if b"\x00" in raw:
        raise BillParseError("CSV 文件包含不允许的二进制内容。")
    decoded = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise BillParseError("CSV 编码不受支持，请使用 UTF-8 或 GB18030。")
    yield from csv.reader(decoded.splitlines())


def _xlsx_rows(path: Path) -> Iterator[list[str]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise BillParseError("XLSX 文件无法读取。") from error
    try:
        worksheet = workbook.worksheets[0]
        for values in worksheet.iter_rows(values_only=True):
            yield ["" if value is None else str(value) for value in values]
    finally:
        workbook.close()


def iter_tabular_rows(path: Path) -> Iterator[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _csv_rows(path)
        return
    if suffix == ".xlsx":
        yield from _xlsx_rows(path)
        return
    raise BillParseError("解析器仅接受 CSV 或 XLSX 文件。")


def normalized_cell(value: str) -> str:
    return " ".join(str(value).replace("\ufeff", "").strip().split())
